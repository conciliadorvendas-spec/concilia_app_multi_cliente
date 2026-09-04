"""
Concilia+ — gerador do relatório Excel.

Recebe o resultado de reconciliation.reconcile() + a config do cliente e
devolve um BytesIO pronto para download. Layout já validado nas versões
anteriores do produto (Relatório Executivo + Divergências Detalhadas).
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NAVY = "1A3A5C"
GREEN = "2ECC71"
RED = "C0392B"
GREY = "7F8C8D"
LIGHT_GREY = "F2F2F2"
FONT_NAME = "Arial"


def _style_header(ws, row, col_start, col_end, color=NAVY):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_workbook(result, config):
    summary = result["summary"]
    sistema_orfaos = result["sistema_orfaos"]
    maquina_orfaos = result["maquina_orfaos"]
    grouped = result["grouped"]
    nome_cliente = config.get("display_name", "Cliente")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ---------- Aba 1: Relatório Executivo ----------
    ws1 = wb.active
    ws1.title = "Relatório Executivo"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 3
    ws1.column_dimensions["B"].width = 42
    ws1.column_dimensions["C"].width = 22

    ws1.merge_cells("B2:C2")
    ws1["B2"] = "Concilia+ — Relatório de Conciliação"
    ws1["B2"].font = Font(name=FONT_NAME, bold=True, size=16, color=NAVY)
    ws1.merge_cells("B3:C3")
    ws1["B3"] = nome_cliente
    ws1["B3"].font = Font(name=FONT_NAME, italic=True, size=10, color="666666")

    row = 5
    ws1.merge_cells(f"B{row}:C{row}")
    ws1[f"B{row}"] = "RESUMO FINANCEIRO"
    ws1[f"B{row}"].font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
    ws1[f"B{row}"].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws1[f"B{row}"].alignment = Alignment(horizontal="center")
    row += 1

    def metric_row(r, label, value, color):
        ws1.merge_cells(f"B{r}:B{r+1}")
        c = ws1[f"B{r}"]
        c.value = label
        c.font = Font(name=FONT_NAME, size=11, color="333333")
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
        ws1.merge_cells(f"C{r}:C{r+1}")
        v = ws1[f"C{r}"]
        v.value = value
        v.font = Font(name=FONT_NAME, bold=True, size=14, color=color)
        v.alignment = Alignment(vertical="center", horizontal="center")
        v.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
        v.number_format = '"R$" #,##0.00'
        for rr in (r, r + 1):
            ws1[f"B{rr}"].border = border
            ws1[f"C{rr}"].border = border
        return r + 2

    row = metric_row(row, "Total de Entradas (vendas conciliadas)", summary["total_entradas"], GREEN)
    row = metric_row(row, "Total em Taxas (Sistema)", summary["total_taxas"], NAVY)
    row = metric_row(row, "Total em Estornos / Chargeback / Cancelamento", summary["total_estornos"], RED)

    row += 1
    ws1.merge_cells(f"B{row}:C{row}")
    ws1[f"B{row}"] = "INDICADORES DE CONCILIAÇÃO"
    ws1[f"B{row}"].font = Font(name=FONT_NAME, bold=True, size=12, color="FFFFFF")
    ws1[f"B{row}"].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    ws1[f"B{row}"].alignment = Alignment(horizontal="center")
    row += 1

    for label, value in [
        ("Parcelas no Sistema (ERP)", summary["n_sistema"]),
        ("Pagamentos na Máquina (Adquirente)", summary["n_maquina"]),
        ("Parcelas conciliadas com sucesso", summary["n_conciliadas"]),
        ("Pagamentos agrupados validados (lote)", summary["n_lotes"]),
        ("Vendas sem correspondência na Máquina", summary["n_orfaos_sistema"]),
        ("Pagamentos sem correspondência no Sistema", summary["n_orfaos_maquina"]),
    ]:
        ws1[f"B{row}"] = label
        ws1[f"B{row}"].font = Font(name=FONT_NAME, size=10, color="333333")
        ws1[f"B{row}"].border = border
        ws1[f"C{row}"] = value
        ws1[f"C{row}"].font = Font(name=FONT_NAME, size=10, bold=True)
        ws1[f"C{row}"].alignment = Alignment(horizontal="center")
        ws1[f"C{row}"].border = border
        row += 1

    # ---------- Aba 2: Divergências Detalhadas ----------
    ws2 = wb.create_sheet("Divergências Detalhadas")
    ws2.sheet_view.showGridLines = False
    for col, w in {"A": 16, "B": 12, "C": 14, "D": 18, "E": 14, "F": 18, "G": 14, "H": 42}.items():
        ws2.column_dimensions[col].width = w

    r = 1
    ws2.merge_cells(f"A{r}:H{r}")
    ws2[f"A{r}"] = f"Divergências Detalhadas — {nome_cliente}"
    ws2[f"A{r}"].font = Font(name=FONT_NAME, bold=True, size=14, color=NAVY)
    r += 2

    # Seção 1
    ws2.merge_cells(f"A{r}:H{r}")
    ws2[f"A{r}"] = "1. Vendas no Sistema SEM correspondência na Máquina"
    ws2[f"A{r}"].font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    ws2[f"A{r}"].fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    r += 1
    for i, h in enumerate(["ID Sistema", "Parcela", "Valor Bruto", "Data Receb.", "Valor Final", "Status", "Forma", "Detalhes"]):
        ws2.cell(row=r, column=1 + i, value=h)
    _style_header(ws2, r, 1, 8)
    r += 1
    if sistema_orfaos.empty:
        ws2.merge_cells(f"A{r}:H{r}")
        ws2[f"A{r}"] = "Nenhuma divergência encontrada nesta categoria."
        ws2[f"A{r}"].font = Font(name=FONT_NAME, italic=True, size=10, color="666666")
        r += 1
    else:
        for _, row_s in sistema_orfaos.iterrows():
            ws2.cell(row=r, column=1, value=str(row_s["id"]))
            ws2.cell(row=r, column=2, value=row_s["parcela"])
            ws2.cell(row=r, column=3, value=float(row_s["valor"]) if pd.notna(row_s["valor"]) else None)
            ws2.cell(row=r, column=4, value=row_s["data_recebimento"].strftime("%d/%m/%Y") if pd.notna(row_s["data_recebimento"]) else "")
            ws2.cell(row=r, column=5, value=float(row_s["vl_final"]))
            ws2.cell(row=r, column=6, value=str(row_s["status"]))
            ws2.cell(row=r, column=7, value=str(row_s.get("forma", "")))
            ws2.cell(row=r, column=8, value="Não localizado em nenhum pagamento da Máquina")
            for c in range(1, 9):
                ws2.cell(row=r, column=c).border = border
                ws2.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10)
            ws2.cell(row=r, column=3).number_format = '"R$" #,##0.00'
            ws2.cell(row=r, column=5).number_format = '"R$" #,##0.00'
            r += 1
    r += 1

    # Seção 2
    ws2.merge_cells(f"A{r}:H{r}")
    ws2[f"A{r}"] = "2. Pagamentos na Máquina SEM correspondência no Sistema"
    ws2[f"A{r}"].font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    ws2[f"A{r}"].fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    r += 1
    for i, h in enumerate(["ID Máquina", "Valor Pago", "Data Pagto.", "Contrato", "Status", "", "", "Detalhes"]):
        ws2.cell(row=r, column=1 + i, value=h)
    _style_header(ws2, r, 1, 8)
    r += 1
    if maquina_orfaos.empty:
        ws2.merge_cells(f"A{r}:H{r}")
        ws2[f"A{r}"] = "Nenhuma divergência encontrada nesta categoria."
        ws2[f"A{r}"].font = Font(name=FONT_NAME, italic=True, size=10, color="666666")
        r += 1
    else:
        for _, row_m in maquina_orfaos.iterrows():
            ws2.cell(row=r, column=1, value=str(row_m.get("id", "")))
            ws2.cell(row=r, column=2, value=float(row_m["valor"]))
            ws2.cell(row=r, column=3, value=row_m["data_pagamento"].strftime("%d/%m/%Y") if pd.notna(row_m["data_pagamento"]) else "")
            ws2.cell(row=r, column=4, value=str(row_m.get("contrato", "")))
            ws2.cell(row=r, column=5, value=str(row_m["status"]))
            ws2.cell(row=r, column=8, value="Não localizado em nenhuma parcela do Sistema")
            for c in range(1, 9):
                ws2.cell(row=r, column=c).border = border
                ws2.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10)
            ws2.cell(row=r, column=2).number_format = '"R$" #,##0.00'
            r += 1
    r += 1

    # Seção 3 — placeholder de layout (o matching atual não deixa "casado com valor diferente")
    ws2.merge_cells(f"A{r}:H{r}")
    ws2[f"A{r}"] = "3. Vendas casadas mas com Valores Divergentes"
    ws2[f"A{r}"].font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    ws2[f"A{r}"].fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    r += 1
    for i, h in enumerate(["ID", "Data", "Valor Sistema", "Valor Máquina", "Diferença", "", "", "Detalhes"]):
        ws2.cell(row=r, column=1 + i, value=h)
    _style_header(ws2, r, 1, 8)
    r += 1
    ws2.merge_cells(f"A{r}:H{r}")
    ws2[f"A{r}"] = "Nenhuma divergência encontrada nesta categoria."
    ws2[f"A{r}"].font = Font(name=FONT_NAME, italic=True, size=10, color="666666")
    r += 2

    # Seção 4 — lotes validados
    ws2.merge_cells(f"A{r}:H{r}")
    ws2[f"A{r}"] = "4. Pagamentos Agrupados pela Máquina — Validados (não são divergência)"
    ws2[f"A{r}"].font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    ws2[f"A{r}"].fill = PatternFill(start_color=GREY, end_color=GREY, fill_type="solid")
    r += 1
    for i, h in enumerate(["ID Máquina", "Valor do Lote", "Data", "ID Sistema", "Parcela", "Valor da Parcela", "", "Observação"]):
        ws2.cell(row=r, column=1 + i, value=h)
    _style_header(ws2, r, 1, 8, color=GREY)
    r += 1
    if not grouped:
        ws2.merge_cells(f"A{r}:H{r}")
        ws2[f"A{r}"] = "Nenhum pagamento agrupado identificado."
        ws2[f"A{r}"].font = Font(name=FONT_NAME, italic=True, size=10, color="666666")
    else:
        for g in grouped:
            start_r = r
            n_parts = len(g["ids_sistema"])
            for i in range(n_parts):
                ws2.cell(row=r, column=4, value=str(g["ids_sistema"][i]))
                ws2.cell(row=r, column=5, value=g["parcelas"][i])
                ws2.cell(row=r, column=6, value=float(g["valores"][i])).number_format = '"R$" #,##0.00'
                for c in range(1, 9):
                    ws2.cell(row=r, column=c).border = border
                    ws2.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10)
                r += 1
            ws2.merge_cells(start_row=start_r, start_column=1, end_row=r - 1, end_column=1)
            ws2.cell(row=start_r, column=1, value=str(g["id_maquina"]))
            ws2.merge_cells(start_row=start_r, start_column=2, end_row=r - 1, end_column=2)
            cell = ws2.cell(row=start_r, column=2, value=float(g["valor_pago"]))
            cell.number_format = '"R$" #,##0.00'
            ws2.merge_cells(start_row=start_r, start_column=3, end_row=r - 1, end_column=3)
            ws2.cell(row=start_r, column=3, value=g["data"].strftime("%d/%m/%Y") if pd.notna(g["data"]) else "")
            ws2.merge_cells(start_row=start_r, start_column=8, end_row=r - 1, end_column=8)
            ws2.cell(row=start_r, column=8, value=f"Soma de {n_parts} parcelas = valor do lote (validado)")
            for cc in (1, 2, 3, 8):
                c = ws2.cell(row=start_r, column=cc)
                c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                c.font = Font(name=FONT_NAME, size=10, bold=(cc in (1, 2)))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
